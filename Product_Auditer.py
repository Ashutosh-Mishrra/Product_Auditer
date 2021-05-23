#!/usr/bin/env python
__author__ = "Ashutosh Mishra"
__credits__ = ["Ashutosh Mishra"]
__code_name__ = "Product Auditer"
__version__ = "1.0"
__maintainer__ = "Ashutosh Mishra"
__status__ = "Production"

import pandas as pd

import requests

import time

from lxml import html

import urllib3

import os

from datetime import datetime

from openpyxl import load_workbook

from selenium import webdriver

from bs4 import BeautifulSoup as bs

 

urllib3.disable_warnings()
 
dir_path = os.path.dirname(os.path.realpath(__file__))

chromedriver = dir_path + '\Input\chromedriver'

#driver = webdriver.Chrome(chromedriver) #...for chrome

 

class Product_Auditer:

 

    def __init__(self):

        self.input=[]

        self.output = {'Link': [], 'Status': [],'Redirect_URL':[]}

 

    def read_input(self):

        input_file = dir_path + r'\Input\Input_file.xlsx'

        wb = load_workbook(input_file)

        ws = wb['Main']

        max_row = ws.max_row

        max_col = ws.max_column

        for row in ws.iter_rows(min_row=0, min_col=0, max_row=max_row, max_col=max_col):

            for cell in row:

                if cell.value:

                    self.input.append(cell.value)

        wb.close()

        print('Reading input....Done')

        print(f'Total-Links:{len(self.input)}')

 
    def main_run(self):

        count = 1

        flag_counter = 1

        for links in self.input:

            if flag_counter == 11:

                print('\nSleep time...Wait few seconds!\n')

                time.sleep(6)

                flag_counter = 1

            print(f'Processing-Link:{count}')

            links = str(links).rstrip()

            ######  First method using requests  ######

            try:

                result = requests.get(links)

                if str(result) == '<Response [200]>':

                    htmldata = result.text

                    contentdata = html.fromstring(htmldata)

                    pname = contentdata.xpath("//a[@class='off-canvas-link btn btn-success']/text()")

                    if pname == []:

                        pname = contentdata.xpath("//div[@id='add-to-cart-stack']/div[2]//text()")

 

                    if pname == []:

                        self.output['Link'].append(links)

                        self.output['Status'].append("Product  Unavailable")

                        self.output['Redirect_URL'].append(result.url)

                    else:

                        self.output['Link'].append(links)

                        self.output['Status'].append("Product Available")

                        self.output['Redirect_URL'].append(result.url)

                else:

                    self.output['Link'].append(links)

                    self.output['Status'].append("Http Error")

                    self.output['Redirect_URL'].append("Error")

            except Exception as error:

                print(f"Error Occured {error}")

            count += 1

            flag_counter += 1

 

    ######   Second method using bs4  #####

    #         driver.get(links)

    #         time.sleep(3)

    #         response = driver.page_source

    #         pd_page_parser = bs(response,'html.parser')

    #         try:

    #             close_button = driver.find_element_by_xpath("//button[@id='ooc-messaging-dismiss']")

    #             if close_button:

    #                 close_button.click()

    #             self.product_finder(pd_page_parser,links)

    #         except:

    #             self.product_finder(pd_page_parser,links)

    #

    #         count += 1

    #     driver.close()

    # def product_finder(self,temp_page_data,temp_links):

    #     pd_page_parser = temp_page_data

    #     links = temp_links

    #     try:

    #         pd_availability_finder = pd_page_parser.findAll('a', {'class': 'off-canvas-link btn btn-success'})

    #         if pd_availability_finder == []:

    #             self.output['Link'].append(links)

    #             self.output['Status'].append("Product Not Available")

    #         else:

    #             self.output['Link'].append(links)

    #             self.output['Status'].append("Product Available")

    #     except Exception as error:

    #         print(f"Error Occured {error}")

    def write_output(self):

        df = pd.DataFrame(self.output)

        now = datetime.now().strftime("_%d_%b_%y_%I_%M_%p")

        writer = pd.ExcelWriter(dir_path + '\Output\NPL_Auditer_Output' + str(now) + '.xlsx')

        df.to_excel(writer, 'Output', index=False)

        writer.save()

if __name__ == '__main__':
   
   start_time = time.time()

   obj = Product_Auditer()

   obj.read_input()

   obj.main_run()

   obj.write_output()

   print(f'\nExecution time: {(time.time()-start_time)/60} mins')
 